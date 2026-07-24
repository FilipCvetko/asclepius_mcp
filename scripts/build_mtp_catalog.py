#!/usr/bin/env python3
"""Build data/mtp_catalog.json from the ZZZS MTP "Seznam s šifrantom, MK, postopki" Excel.

The `seznam MP` sheet is the authoritative medical-device (medicinsko-tehnični pripomočki) list.
Per device it carries the prescribing rules a clinician needs — most importantly the **indications**
("BOLEZEN / ZDRAVSTVENO STANJE in DRUGI POGOJI"), plus who may prescribe, whether an odločba IOZ is
required, duration, naročilnica rules and the price standard. (The previous version of this script
dropped exactly the indications/naročilnica/price columns — they are the whole point.)

Source = the latest file fetched by scripts/fetch_mtp.py (data/mtp/latest.json); falls back to the
newest inner sheet of the e-gradiva-crawled zip if that's absent. Output is loaded by mtp.py.

The sheet is hierarchical: category/section rows (ŠIFRA holds descriptive text, IME empty) introduce
groups of device rows; we carry the current category path onto each device.
"""
import io
import json
import re
import sys
import zipfile
from pathlib import Path

import openpyxl

# Resolve the sifrant catalog via paths.data_path (writable volume wins over the image-baked copy)
# so the periodic refresh, which rebuilds the sifrant catalog onto the volume BEFORE rebuilding MTP,
# feeds the freshest tables into the enrichment. paths.py lives at the repo root (one up from scripts/).
sys.path.insert(0, str(Path(__file__).parent.parent))
try:
    from paths import data_path as _data_path
except Exception:
    _data_path = None

DATA = Path(__file__).parent.parent / "data"
MTP_DIR = DATA / "mtp"
ZIP_FALLBACK = DATA / "egradiva_files" / "DFDC914987E44E2AC1257353003EC73A.other"
FALLBACK_URL = ("https://www.zzzs.si/?id=126&detail=DFDC914987E44E2AC1257353003EC73A")
OUT = DATA / "mtp_catalog.json"

# Companion structured files (crawled snapshots in egradiva_files):
#  - mutual-exclusion list: devices that may NOT be prescribed together (Pravila OZZ).
#  - timska obravnava: devices requiring a team assessment before prescription.
# Companion files are baked into the image at data/mtp_companions/ (small; rarely change) so the
# periodic refresh can rebuild MTP with exclusions without the dockerignored egradiva_files/.
# Override with MTP_EXCL_FILE / MTP_TIMSKA_FILE to point at fresher copies on the volume.
import os as _os
EXCL_FILE = Path(_os.environ.get("MTP_EXCL_FILE", DATA / "mtp_companions" / "exclusions.xlsx"))
EXCL_URL = ("https://www.zzzs.si/?id=126")  # Zbirka podatkov za MP (e-gradiva)
TIMSKA_FILE = Path(_os.environ.get("MTP_TIMSKA_FILE", DATA / "mtp_companions" / "timska.xlsx"))


def _norm_code(c) -> str:
    """Normalize an MP code so the exclusion list ('131') matches the catalog ('0131')."""
    s = str(c or "").strip()
    return s.zfill(4) if s.isdigit() else s


def load_exclusions():
    """Return {norm_code: [{sifra, naziv}, …]} of mutually-exclusive devices (symmetric)."""
    if not EXCL_FILE.exists():
        print("WARNING: MP exclusion file missing; skipping izkljucuje")
        return {}
    import io
    wb = openpyxl.load_workbook(io.BytesIO(EXCL_FILE.read_bytes()), read_only=True, data_only=True)
    ws = wb["List1"] if "List1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = next((i for i, r in enumerate(rows)
                  if r and str(r[0] or "").strip().upper() == "ŠIFRA MP"), 2)
    pairs, naziv = {}, {}
    for r in rows[hdr_i + 1:]:
        if not r or not str(r[0] or "").strip().isdigit() or len(r) < 4:
            continue
        a, b = _norm_code(r[0]), _norm_code(r[2])
        if a == b:                                   # ignore self-references
            continue
        naziv[a] = _clean(r[1]); naziv[b] = _clean(r[3])
        pairs.setdefault(a, set()).add(b)
        pairs.setdefault(b, set()).add(a)            # symmetric
    return {c: [{"sifra": x, "naziv": naziv.get(x, "")} for x in sorted(s)]
            for c, s in pairs.items()}


def load_timska():
    """Return the set of norm codes requiring timska obravnava (team assessment)."""
    if not TIMSKA_FILE.exists():
        return set()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(TIMSKA_FILE.read_bytes()), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return {_norm_code(r[0]) for r in ws.iter_rows(values_only=True)
            if r and str(r[0] or "").strip() and str(r[0]).strip()[0].isdigit()}

# Ordered (header-substring, field). First match wins per column, so specific/disambiguating
# patterns come BEFORE generic ones (e.g. "PODALJŠANJE" before "IZPOSOJA"/"OBNOVLJIVA", and
# "PREDPIS NAROČILNICE" before the bare naročilnica columns).
COLS = [
    ("ŠIFRA", "sifra"),
    ("IME PRIPOMOČKA", "ime"),
    ("PRISTOJNOST", "pristojnost"),
    ("ODLOČBA", "odlocba"),
    ("MATERIALNI STROŠEK", "materialni_strosek"),
    ("PODALJŠANJE", "podaljsanje_izposoje"),
    ("OBNOVLJIVA NAROČILNICA", "obnovljiva_narocilnica"),
    ("PREDPIS NAROČILNICE", "predpis_narocilnice"),
    ("IZPOSOJA", "izposoja"),
    ("POPRAVILA", "popravila"),
    ("PRILAGODITVE", "prilagoditve"),
    ("DOBA TRAJANJA", "doba_trajanja"),
    ("BOLEZEN", "indikacije"),
    ("ZDRAVSTVENO STANJE", "indikacije"),
    ("CENOVNI", "cenovni_standard"),
]


def _clean(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v not in (None, "") else ""


def _load_source():
    """Return (workbook_bytes, source_filename, datum, source_url) for the newest list."""
    latest = MTP_DIR / "latest.json"
    if latest.exists():
        meta = json.loads(latest.read_text())
        xlsx = MTP_DIR / meta["xlsx_file"]
        if xlsx.exists():
            return xlsx.read_bytes(), meta["xlsx_file"], meta["datum"], meta.get("source_url", "")
    # fallback: newest inner xlsx of the crawled zip (stale snapshot)
    print("WARNING: data/mtp/latest.json missing — using bundled zip fallback "
          "(run scripts/fetch_mtp.py for the current file)")
    zf = zipfile.ZipFile(ZIP_FALLBACK)

    def dt(n):
        m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", n)
        return (int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else (0, 0, 0)
    inner = sorted((n for n in zf.namelist() if n.lower().endswith(".xlsx")), key=dt)[-1]
    y, mo, d = dt(inner)
    return zf.read(inner), inner, f"{y:04d}-{mo:02d}-{d:02d}", FALLBACK_URL


def _map_columns(header):
    field_at = {}
    for idx, cell in enumerate(header):
        h = (str(cell).upper() if cell else "")
        for pat, fld in COLS:
            if pat in h:
                field_at[idx] = fld
                break
    return field_at


# ── Structured enrichment from the ZZZS "Vsi šifranti" catalog ─────────────────────────────────
# The device Excel lacks the authoritative prescriber category, period×quantity and EHR grouping —
# those live in the "Vsi šifranti" MP tables (built by build_sifrant_catalog.py). We JOIN them here,
# at build time, by the 4-digit device code (SifraVrsMTP == device sifra), so the runtime tool
# (mtp.py) stays dependency-free. Requires build_sifrant_catalog.py's KEEP_ROWS retention of the MP
# link tables; if the catalog is absent/old the loader returns None and only base fields are kept.
_PEER_CAP = 40

# Durability modes (VrsteMTP.OpisDobeTrajanja) that denote a one-off / lifespan device rather than a
# recurring per-period quantity — the renderer must NOT invent a per-day count for these.
_DURABLE_MODES = {"TRAJNOSTNA DOBA", "ŽIVLJENJSKA DOBA", "ENKRATNA PRAVICA"}


def _cur(rows):
    """Rows still valid (no VeljaDo end date) — mirrors sifranti.py's current-only convention."""
    return [r for r in rows if not r.get("VeljaDo")]


def load_sifrant_enrichment():
    """Load + index the MP tables from the sifrant catalog. Returns the index dict, or None when the
    catalog (or its MP link rows) is unavailable → enrichment is skipped, base fields kept."""
    p = _data_path("sifrant_catalog.json") if _data_path else (DATA / "sifrant_catalog.json")
    if not p or not p.exists():
        print("WARNING: sifrant_catalog.json not found — MTP enrichment skipped")
        return None
    sif = json.loads(p.read_text(encoding="utf-8")).get("sifranti", {})

    def rows(oz):
        return (sif.get(oz) or {}).get("zapisi", []) or []

    if not rows("K42") and not rows("K38.3"):
        print("WARNING: sifrant catalog lacks MP link tables (rebuild build_sifrant_catalog.py) "
              "— MTP enrichment skipped")
        return None

    vrste = {_norm_code(r["SifraVrsMTP"]): r for r in rows("15.40") if r.get("SifraVrsMTP")}
    doba_by_code = {r["SifraTrajDobMTP"]: r for r in rows("35") if r.get("SifraTrajDobMTP")}

    # device -> [trajnostna doba rows] (current links only)
    k42 = {}
    for r in _cur(rows("K42")):
        c, td = _norm_code(r.get("SifVrsMTP")), doba_by_code.get(r.get("SifraTrajDobMTP"))
        if c and td:
            k42.setdefault(c, []).append(td)

    # group descriptor names
    sku_naziv = {r.get("SifraSkuMTP"): r.get("NazivSkuMTP", "") for r in rows("34.2")}
    pod1_naziv = {(r.get("SifraSkuMTP"), r.get("SifPodSkuMTP1")): r.get("NazivPodSkuMTP1", "")
                  for r in rows("34.4")}
    pod2_naziv = {(r.get("SifraSkuMTP"), r.get("SifPodSkuMTP1"), r.get("SifPodSkuMTP2")):
                  r.get("NazivPodSkuMTP2", "") for r in rows("34.5")}
    podgruca = {r["SifraPodgruceVrsMP"]: r for r in rows("60") if r.get("SifraPodgruceVrsMP")}

    # device -> list of group tuples (sku, pod1|None, pod2|None), deepest level it is filed under
    membership = {}
    for r in _cur(rows("K38.3")):
        membership.setdefault(_norm_code(r.get("SifraVrsMTP")), []).append(
            (r.get("SifraSkuMTP"), r.get("SifPodSkuMTP1"), r.get("SifPodSkuMTP2")))
    for r in _cur(rows("K38.2")):
        membership.setdefault(_norm_code(r.get("SifraVrsMTP")), []).append(
            (r.get("SifraSkuMTP"), r.get("SifPodSkuMTP1"), None))
    for r in _cur(rows("K38.1")):
        membership.setdefault(_norm_code(r.get("SifraVrsMTP")), []).append(
            (r.get("SifraSkuMTP"), None, None))
    by_tuple = {}                                     # group tuple -> {device codes}
    for c, tups in membership.items():
        for t in tups:
            by_tuple.setdefault(t, set()).add(c)

    # podgruča (finer family) -> {device codes}, from VrsteMTP.SifraPodgruceVrsMP
    by_podgruca = {}
    for c, r in vrste.items():
        pg = r.get("SifraPodgruceVrsMP")
        if pg and pg != "0":
            by_podgruca.setdefault(pg, set()).add(c)

    # article systems: device -> [{sistem, sistem_naziv, artikel, artikel_naziv}] (current links)
    art_naziv = {r.get("SifraArtiklaZZZS"): r.get("NazivArtiklaZZZS", "") for r in rows("64")}
    sist_naziv = {r.get("SifraSistemaZZZS"): r.get("NazivSistemaZZZS", "") for r in rows("63")}
    systems = {}
    for r in _cur(rows("K45.1")) + _cur(rows("K45.2")):
        c = _norm_code(r.get("SifraVrsMTP"))
        if not c:
            continue
        systems.setdefault(c, []).append(
            {"sistem": r.get("SifraSistemaZZZS"), "sistem_naziv": sist_naziv.get(r.get("SifraSistemaZZZS"), ""),
             "artikel": r.get("SifraArtiklaZZZS"), "artikel_naziv": art_naziv.get(r.get("SifraArtiklaZZZS"), "")})

    return {"vrste": vrste, "k42": k42, "sku_naziv": sku_naziv, "pod1_naziv": pod1_naziv,
            "pod2_naziv": pod2_naziv, "podgruca": podgruca, "membership": membership,
            "by_tuple": by_tuple, "by_podgruca": by_podgruca, "systems": systems}


def _enrich_device(d, enr, name_by_code, catalog_codes):
    """Attach prescriber / quantity_periods / group / group_peers / systems to one device dict."""
    nc = _norm_code(d.get("sifra"))
    vr = enr["vrste"].get(nc)
    if not vr:
        return                                        # not in VrsteMTP → keep base fields only

    # (1) WHO may prescribe — OpisPoobZdra verbatim is authoritative; flags are derived, not invented
    opis = vr.get("OpisPoobZdra", "") or ""
    low = opis.lower()
    d["prescriber"] = {
        "opis": opis,
        "nurse": "dms" in low,
        "personal": "osebni" in low or "splošni osebni" in low,
        "specialist": any(k in low for k in ("specialist", "diabetolog", "ginekolog")),
        "odlocba_ioz": vr.get("OzPotOdlIOZ", ""),
        "predpis_dovoljen": vr.get("OzDovolPredpisVrsMP", ""),
    }

    # (2) PERIOD + QUANTITY — from the linked trajnostna-doba rows; mode from OpisDobeTrajanja
    d["nacin_dobe"] = vr.get("OpisDobeTrajanja", "")
    for src, dst in (("StKosovVPakiranju", "st_kosov_pakiranje"),
                     ("OzDvojnaKolic", "dvojna_kolicina"), ("OpisRazKolic", "razpon_kolicine")):
        if vr.get(src):
            d[dst] = vr.get(src)
    qp = []
    for td in enr["k42"].get(nc, []):
        qp.append({"opis": td.get("OpisTrajDobMTP", ""),
                   "trajanje": td.get("TrajDoba", ""), "trajanje_enota": td.get("EnotMerTrajDob", ""),
                   "max_kolicina": td.get("MaxDovKol", ""), "kolicina_enota": td.get("EnotaMerMaxDovKol", ""),
                   "na_dan_min": td.get("MinKolicNaDan", ""), "na_dan_max": td.get("MaxKolicNaDan", ""),
                   "starost_od": td.get("StarOseOd", ""), "starost_do": td.get("StarOseDo", "")})
    if qp:
        d["quantity_periods"] = qp
    d["doba_mode"] = "durable" if (d["nacin_dobe"] in _DURABLE_MODES or not qp) else "periodic"

    # (3) GROUP + peers prescribable alongside (same group MINUS mutually-exclusive; there is NO
    # positive "prescribe-together" table in the source — this is a same-group-minus-exclusions view)
    pg = vr.get("SifraPodgruceVrsMP")
    tups = enr["membership"].get(nc, [])
    group = {}
    if tups:
        skus = sorted({t[0] for t in tups if t[0]})
        group["skupine"] = [{"sifra": s, "naziv": enr["sku_naziv"].get(s, "")} for s in skus]
        subs = []
        for (sku, p1, p2) in tups:
            item = {"skupina": sku}
            if p1:
                item["pod1"], item["pod1_naziv"] = p1, enr["pod1_naziv"].get((sku, p1), "")
            if p2:
                item["pod2"], item["pod2_naziv"] = p2, enr["pod2_naziv"].get((sku, p1, p2), "")
            subs.append(item)
        group["podskupine"] = subs
    if pg and pg != "0":
        g = enr["podgruca"].get(pg, {})
        group.update({"podgruca": pg, "podgruca_naziv": g.get("OpisPodgruceVrsMP", ""),
                      "gruca": g.get("SifraGruceVrsMP", ""), "gruca_naziv": g.get("OpisGruceVrsMP", "")})
    if group:
        d["group"] = group

    # peers: finest available grouping — podgruča if present, else union of the K38 group tuples;
    # restricted to devices actually in the current prescribable catalog.
    if pg and pg != "0":
        peers = set(enr["by_podgruca"].get(pg, ()))
    else:
        peers = set()
        for t in tups:
            peers |= enr["by_tuple"].get(t, set())
    peers = (peers & catalog_codes) - {nc}
    if peers:
        excl = {_norm_code(e.get("sifra")) for e in (d.get("izkljucuje") or [])}
        ordered = sorted(peers)
        d["group_peers"] = [{"sifra": c, "ime": name_by_code.get(c, ""), "excluded": c in excl}
                            for c in ordered[:_PEER_CAP]]
        if len(ordered) > _PEER_CAP:
            d["group_peers_overflow"] = len(ordered) - _PEER_CAP

    # (optional) article systems (components that form one treatment system)
    if enr["systems"].get(nc):
        d["systems"] = enr["systems"][nc]


def main():
    blob, src_name, datum, src_url = _load_source()
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb["seznam MP"] if "seznam MP" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    hdr_i = next(i for i, r in enumerate(rows)
                 if r and any(str(c).strip().upper() == "ŠIFRA" for c in r if c))
    field_at = _map_columns(rows[hdr_i])

    devices, categories, cat_path = [], 0, []
    for r in rows[hdr_i + 1:]:
        if not r or r[0] is None or not str(r[0]).strip():
            continue
        vals = {field_at[i]: _clean(r[i]) for i in field_at if i < len(r)}
        sifra, ime = vals.get("sifra", ""), vals.get("ime", "")
        has_data = any(vals.get(f) for f in ("ime", "pristojnost", "indikacije", "materialni_strosek"))
        if ime and has_data:                      # a real device row
            devices.append({**vals, "kategorija": "; ".join(cat_path),
                            "source": src_name, "source_url": src_url, "datum": datum})
        else:                                     # a category / section header
            if re.match(r"^\d+\.", sifra):
                cat_path = [sifra]
            elif cat_path:
                cat_path = cat_path[:1] + [sifra]
            else:
                cat_path = [sifra]
            categories += 1

    # Attach mutually-exclusive devices + timska-obravnava flag (companion structured files)
    exclusions = load_exclusions()
    timska = load_timska()
    n_excl = 0
    for d in devices:
        nc = _norm_code(d.get("sifra"))
        if nc in exclusions:
            d["izkljucuje"] = exclusions[nc]
            n_excl += 1
        if nc in timska:
            d["timska_obravnava"] = "DA"

    # Attach structured enrichment (prescriber / period+quantity / group peers) from the ZZZS
    # "Vsi šifranti" catalog. Runs AFTER exclusions so peers can be flagged combinable-vs-not.
    enr = load_sifrant_enrichment()
    n_enr = n_qty = n_group = 0
    if enr:
        name_by_code = {_norm_code(d["sifra"]): d.get("ime", "") for d in devices if d.get("sifra")}
        catalog_codes = set(name_by_code)
        for d in devices:
            _enrich_device(d, enr, name_by_code, catalog_codes)
            if d.get("prescriber"):
                n_enr += 1
            if d.get("quantity_periods"):
                n_qty += 1
            if d.get("group_peers"):
                n_group += 1

    out = {"source": src_name, "source_url": src_url, "datum": datum,
           "count": len(devices), "fields": [f for _, f in COLS],
           "exclusion_pairs": sum(len(v) for v in exclusions.values()) // 2,
           "enriched": n_enr, "devices": devices}
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"  exclusions: {n_excl} devices flagged; timska obravnava: "
          f"{sum(1 for d in devices if d.get('timska_obravnava'))} devices")
    print(f"  enrichment (Vsi šifranti): {n_enr}/{len(devices)} prescriber, "
          f"{n_qty} with quantity_periods, {n_group} with group peers")
    with_ind = sum(1 for d in devices if d.get("indikacije"))
    print(f"MTP catalog: {len(devices)} devices ({with_ind} with indications), "
          f"{categories} category rows, datum {datum}")
    print(f"  source: {src_name}")
    print(f"  -> {OUT}  ({OUT.stat().st_size/1024:.0f} KB)")
    for d in devices:
        if "BERGLA" in d.get("ime", "").upper():
            print(f"  e.g. {d['sifra']} {d['ime']}: predpiše={d.get('pristojnost')} | "
                  f"odločba={d.get('odlocba')} | indik={d.get('indikacije','')[:60]}…")
            break


if __name__ == "__main__":
    main()
