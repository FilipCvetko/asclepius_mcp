"""MKB-10 / ICD-10 diagnostic codes module - NIJZ data, cached locally."""

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from cache import is_cache_fresh, read_cache, write_cache

# Configuration
DATA_DIR = Path(__file__).parent / "data"
ICD10_DATA_FILE = DATA_DIR / "icd10_codes.json"
CACHE_NAME = "icd10"
CACHE_TTL_HOURS = 720  # 30 days - ICD codes rarely change
NIJZ_XLSX_URL = "https://nijz.si/wp-content/uploads/2024/12/MKB-10-AM-v11_2023.xlsx"
CODE_PATTERN = re.compile(r"^[A-Z]\d{2}(\.\d+)?$", re.IGNORECASE)

# Module-level state
ICD10_CODES: List[Dict[str, Any]] = []
VECTORIZER: Optional[TfidfVectorizer] = None
VECTORS: Optional[Any] = None


def _download_and_parse_mkb10() -> List[Dict[str, Any]]:
    """Download XLSX from NIJZ, parse sheet 'MKB-10-AM v11', return ICD-10 records.

    XLSX structure (sheet 'MKB-10-AM v11'):
        Col 0: POGLAVJE, Col 1: KATEGORIJA, Col 4: SKLOP / 1. NIVO,
        Col 7: RAVEN, Col 8: KODA, Col 9: SLOVENSKI NAZIV
    """
    try:
        import openpyxl
        import requests
    except ImportError as e:
        print(f"Missing dependency for ICD-10 download: {e}")
        return []

    import tempfile

    print("Downloading MKB-10 data from NIJZ...")
    try:
        req_headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        }
        response = requests.get(NIJZ_XLSX_URL, headers=req_headers, timeout=60)
        response.raise_for_status()
    except Exception as e:
        print(f"Failed to download MKB-10 XLSX: {e}")
        return []

    codes = []
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(response.content)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

        # Use the main data sheet
        sheet_name = "MKB-10-AM v11"
        if sheet_name not in wb.sheetnames:
            # Fallback: try second sheet or active
            sheet_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]
        ws = wb[sheet_name]

        # Skip header row (row 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 10:
                continue

            code = str(row[8] or "").strip()   # KODA
            desc_sl = str(row[9] or "").strip()  # SLOVENSKI NAZIV
            category_raw = str(row[4] or "").strip()  # SKLOP / 1. NIVO

            if not code or not desc_sl:
                continue

            code = code.upper().replace(" ", "")

            # Format category from SKLOP (e.g. "A00A09" -> "A00-A09")
            if len(category_raw) == 6 and category_raw[0].isalpha():
                category = f"{category_raw[:3]}-{category_raw[3:]}"
            else:
                category = category_raw

            codes.append({
                "code": code,
                "description_sl": desc_sl,
                "description_en": "",
                "category": category,
            })

        wb.close()
        Path(tmp_path).unlink(missing_ok=True)

    except Exception as e:
        print(f"Error parsing MKB-10 XLSX: {e}")
        return []

    print(f"Parsed {len(codes)} ICD-10 codes from NIJZ")
    return codes


def _load_bundled_data() -> List[Dict[str, Any]]:
    """Load bundled ICD-10 data from data/icd10_codes.json."""
    if ICD10_DATA_FILE.exists():
        try:
            with open(ICD10_DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading bundled ICD-10 data: {e}")
    return []


def _save_bundled_data(codes: List[Dict[str, Any]]) -> None:
    """Save parsed ICD-10 data to data/icd10_codes.json."""
    DATA_DIR.mkdir(exist_ok=True)
    try:
        with open(ICD10_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(codes, f, ensure_ascii=False, indent=2)
        print(f"Saved {len(codes)} ICD-10 codes to {ICD10_DATA_FILE}")
    except Exception as e:
        print(f"Error saving ICD-10 data: {e}")


def _build_index(codes: List[Dict[str, Any]]) -> None:
    """Build TF-IDF index on code + descriptions."""
    global VECTORIZER, VECTORS

    if not codes:
        return

    corpus = [
        f"{c['code']} {c.get('description_sl', '')} {c.get('description_en', '')}"
        for c in codes
    ]
    VECTORIZER = TfidfVectorizer()
    VECTORIZER.fit(corpus)
    VECTORS = VECTORIZER.transform(corpus)


def initialize_icd10() -> None:
    """Initialize ICD-10 module: load data, build search index."""
    global ICD10_CODES

    print("Loading ICD-10/MKB-10 codes...")

    # Try fresh cache first
    if is_cache_fresh(CACHE_NAME, max_age_hours=CACHE_TTL_HOURS):
        cached = read_cache(CACHE_NAME)
        if cached:
            ICD10_CODES = cached
            _build_index(ICD10_CODES)
            print(f"Loaded {len(ICD10_CODES)} ICD-10 codes from cache")
            return

    # Try downloading fresh data
    try:
        codes = _download_and_parse_mkb10()
        if codes:
            ICD10_CODES = codes
            write_cache(CACHE_NAME, codes)
            _save_bundled_data(codes)
            _build_index(ICD10_CODES)
            print(f"Loaded {len(ICD10_CODES)} ICD-10 codes from NIJZ")
            return
    except Exception as e:
        print(f"Warning: Failed to download ICD-10 data: {e}")

    # Fall back to stale cache
    cached = read_cache(CACHE_NAME)
    if cached:
        ICD10_CODES = cached
        _build_index(ICD10_CODES)
        print(f"Loaded {len(ICD10_CODES)} ICD-10 codes from stale cache")
        return

    # Fall back to bundled data
    bundled = _load_bundled_data()
    if bundled:
        ICD10_CODES = bundled
        write_cache(CACHE_NAME, bundled)
        _build_index(ICD10_CODES)
        print(f"Loaded {len(ICD10_CODES)} ICD-10 codes from bundled data")
        return

    print("Warning: No ICD-10 data available")


def _get_icd10(query: str) -> Dict[str, Any]:
    """Bidirectional ICD-10 lookup: code→description or description→codes."""
    if not ICD10_CODES:
        # Try lazy refresh
        initialize_icd10()
        if not ICD10_CODES:
            return {"found": False, "error": "No ICD-10 data available"}

    query = query.strip()
    if not query:
        return {"found": False, "error": "Query must not be empty"}

    # Check if query looks like an ICD-10 code
    if CODE_PATTERN.match(query):
        return _lookup_by_code(query.upper())
    else:
        return _search_by_description(query)


def _lookup_by_code(code: str) -> Dict[str, Any]:
    """Exact code lookup, then prefix match."""
    results = []

    # Exact match first
    for c in ICD10_CODES:
        if c["code"] == code:
            results.append({**c, "match_type": "exact"})
            break

    # Prefix match (e.g. "J06" matches "J06.0", "J06.9", etc.)
    for c in ICD10_CODES:
        if c["code"].startswith(code) and c["code"] != code:
            results.append({**c, "match_type": "prefix"})

    if not results:
        return {"found": False, "error": f"No ICD-10 code matching '{code}'", "query": code}

    return {
        "found": True,
        "count": len(results),
        "query": code,
        "results": results[:20],
    }


def _search_by_description(query: str) -> Dict[str, Any]:
    """TF-IDF search across descriptions."""
    if VECTORIZER is None or VECTORS is None:
        return {"found": False, "error": "Search index not initialized"}

    query_vector = VECTORIZER.transform([query])
    similarities = cosine_similarity(query_vector, VECTORS).flatten()

    results = []
    for idx, score in enumerate(similarities):
        if score > 0.15:
            results.append({
                **ICD10_CODES[idx],
                "confidence": round(float(score), 4),
            })

    results.sort(key=lambda x: x["confidence"], reverse=True)

    if not results:
        return {"found": False, "error": "No matching ICD-10 codes found", "query": query}

    return {
        "found": True,
        "count": len(results[:20]),
        "query": query,
        "results": results[:20],
    }


if __name__ == "__main__":
    print("Testing ICD-10 module...")
    initialize_icd10()

    # Test code lookup
    print("\n--- Code lookup: J06.9 ---")
    result = _get_icd10("J06.9")
    print(json.dumps(result, ensure_ascii=False, indent=2)[:500])

    # Test description search
    print("\n--- Description search: glavobol ---")
    result = _get_icd10("glavobol")
    print(json.dumps(result, ensure_ascii=False, indent=2)[:500])
